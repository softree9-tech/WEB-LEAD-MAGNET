import io
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
import openpyxl
from reportlab.graphics.shapes import Drawing, Rect

BRAND_DARK = colors.HexColor('#1E293B')

def sanitize_text(text):
    if text is None:
        return ''
    return str(text).replace('\r', ' ').replace('\n', ' ').strip()

import datetime

def create_score_bar(score, max_score=100, width=200, height=12):
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=colors.HexColor('#E2E8F0'), strokeColor=None, rx=height/2, ry=height/2))
    fill_width = (score / max_score) * width
    d.add(Rect(0, 0, fill_width, height, fillColor=colors.HexColor('#FF6B35'), strokeColor=None, rx=height/2, ry=height/2))
    return d

def generate_pdf_report(report_data: dict) -> bytes:
    buffer = io.BytesIO()
    
    BRAND_ORANGE = colors.HexColor('#FF6B35')
    BRAND_DARK = colors.HexColor('#1E293B')
    BRAND_GRAY = colors.HexColor('#64748B')
    BRAND_LIGHT_BG = colors.HexColor('#F8FAFC')
    
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=40, leftMargin=40,
        topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=22, leading=26,
        textColor=BRAND_DARK, spaceAfter=20
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle', parent=styles['Heading2'],
        fontName='Helvetica-Bold', fontSize=14, leading=18,
        textColor=BRAND_ORANGE, spaceAfter=12, spaceBefore=20
    )
    
    normal_text = ParagraphStyle(
        'NormalText', parent=styles['Normal'],
        fontName='Helvetica', fontSize=10, leading=14,
        textColor=BRAND_GRAY
    )
    
    bold_text = ParagraphStyle(
        'BoldText', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=10, leading=14,
        textColor=BRAND_DARK
    )

    card_title_style = ParagraphStyle(
        'CardTitle', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=11, leading=14,
        textColor=BRAND_ORANGE, alignment=1
    )
    
    card_value_style = ParagraphStyle(
        'CardValue', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=18, leading=22,
        textColor=BRAND_DARK, alignment=1
    )
    
    elements = []
    
    # Header
    date_str = datetime.datetime.now().strftime("%B %d, %Y")
    header_data = [
        [Paragraph("<b>Softree Technology</b>", ParagraphStyle('H1', fontName='Helvetica-Bold', fontSize=16, textColor=BRAND_ORANGE)),
         Paragraph(f"<font color='#64748B'>{date_str}</font>", ParagraphStyle('H2', alignment=2, fontName='Helvetica', fontSize=10))]
    ]
    header_table = Table(header_data, colWidths=[250, 250])
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elements.append(header_table)
    elements.append(Spacer(1, 15))
    
    website = report_data.get('website', 'Domain')
    elements.append(Paragraph("Executive Audit Report", title_style))
    elements.append(Paragraph(f"Comprehensive Website Intelligence & Conversion Analysis for <b>{website}</b>", normal_text))
    elements.append(Spacer(1, 15))
    
    # Calculate UX Score
    consistencyVal = 90 if report_data.get('design') == 'Modern' else 60
    flowVal = 80 if report_data.get('message') == 'Clear' else 50
    mobileVal = 80 if report_data.get('seo_mobile') else 30
    engagementVal = 90 if report_data.get('cta') == 'Strong' else 40
    uxScore = round((consistencyVal + flowVal + mobileVal + engagementVal) / 4)
    
    seo_score = int(report_data.get('seo_score', 0))
    aeo_score = int(report_data.get('aeo_score', 0))
    
    # Metric Cards
    metrics_data = [
        [Paragraph("UX & Conversion Score", card_title_style), Paragraph("SEO Performance", card_title_style), Paragraph("AI Visibility (AEO)", card_title_style)],
        [Paragraph(f"{uxScore}/100", card_value_style), Paragraph(f"{seo_score}/100", card_value_style), Paragraph(f"{aeo_score}/100", card_value_style)]
    ]
    t = Table(metrics_data, colWidths=[170, 170, 170])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), BRAND_LIGHT_BG),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 15),
        ('BOTTOMPADDING', (0,0), (-1,-1), 15),
        ('GRID', (0,0), (-1,-1), 1, colors.white),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 15))
    
    # Executive Summary & Competitor Battle Card
    elements.append(Paragraph("Executive Presence & Competitor Benchmark", subtitle_style))
    elements.append(Paragraph(sanitize_text(report_data.get('executive_summary', 'Analysis completed successfully.')), normal_text))
    elements.append(Spacer(1, 10))
    
    # Battle Card Table
    if report_data.get('battle_data'):
        comp_data = report_data.get('competitor_data', {})
        elements.append(Paragraph(f"Competitor Comparison: vs {sanitize_text(comp_data.get('website', 'Competitor'))}", bold_text))
        bc_data = [
            [Paragraph("Metric", bold_text), Paragraph("Primary", bold_text), Paragraph("Competitor", bold_text)],
            [Paragraph("SEO Score", normal_text), Paragraph(f"{seo_score}", normal_text), Paragraph(f"{comp_data.get('seo_score', 'N/A')}", normal_text)],
            [Paragraph("AI Visibility", normal_text), Paragraph(f"{aeo_score}", normal_text), Paragraph(f"{comp_data.get('aeo_score', 'N/A')}", normal_text)],
            [Paragraph("Performance", normal_text), Paragraph(f"{report_data.get('lighthouse_performance', 0)}", normal_text), Paragraph(f"{comp_data.get('lighthouse_performance', 'N/A')}", normal_text)]
        ]
        bct = Table(bc_data, colWidths=[150, 150, 150])
        bct.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('BACKGROUND', (0,0), (-1,0), BRAND_LIGHT_BG),
            ('PADDING', (0,0), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(bct)
        elements.append(Spacer(1, 10))
        
        bd = report_data.get('battle_data', {})
        elements.append(Paragraph(f"<b>Overall Advantage:</b> {sanitize_text(bd.get('overall_winner', 'N/A'))}", bold_text))
        elements.append(Paragraph(sanitize_text(bd.get('ai_verdict', '')), normal_text))
        elements.append(Spacer(1, 15))
    
    elements.append(Paragraph("Business Risk Insight", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('business_risk_insight', '')), normal_text))
    elements.append(Spacer(1, 15))
    
    # UX Scorecard Breakdown
    elements.append(Paragraph("Rebranding UX Scorecard", subtitle_style))
    
    ux_breakdown_data = [
        [Paragraph("Consistency", bold_text), Paragraph("9/10" if report_data.get('design') == 'Modern' else "6/10", normal_text),
         Paragraph("Visual Flow", bold_text), Paragraph("8/10" if report_data.get('message') == 'Clear' else "5/10", normal_text)],
        [Paragraph("Mobile UX", bold_text), Paragraph("8/10" if report_data.get('seo_mobile') else "3/10", normal_text),
         Paragraph("User Engagement", bold_text), Paragraph("9/10" if report_data.get('cta') == 'Strong' else "4/10", normal_text)]
    ]
    uxt_bd = Table(ux_breakdown_data, colWidths=[125, 125, 125, 125])
    uxt_bd.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(uxt_bd)
    elements.append(Spacer(1, 10))
    
    # Mobile UX Reality Insight
    elements.append(Paragraph("Mobile Walkthrough Insight", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('mobile_conversion_recommendation', 'Optimize mobile viewport for seamless conversions.')), normal_text))
    elements.append(Spacer(1, 15))
    
    # UX & Conversion Metrics
    elements.append(Paragraph("UX & Conversion Analysis", subtitle_style))
    
    ux_table_data = [
        [Paragraph("Mobile UX Rating", bold_text), Paragraph(sanitize_text(report_data.get('mobile_ux_rating', 'Average')), normal_text)],
        [Paragraph("Conversion Risk", bold_text), Paragraph(sanitize_text(report_data.get('mobile_conversion_risk', 'Moderate')), normal_text)],
        [Paragraph("Readiness Level", bold_text), Paragraph(sanitize_text(report_data.get('conversion_readiness_level', 'Low')), normal_text)],
        [Paragraph("Revenue Leak Severity", bold_text), Paragraph(sanitize_text(report_data.get('revenue_leak_severity', 'Low')), normal_text)],
    ]
    uxt = Table(ux_table_data, colWidths=[150, 350])
    uxt.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0,0), (0,-1), BRAND_LIGHT_BG),
        ('PADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(uxt)
    elements.append(Spacer(1, 10))
    
    # Detailed Conversion Opportunities
    elements.append(Paragraph("Missing Conversion Paths", bold_text))
    missing_opps = report_data.get('missing_opportunities_list', [])
    if missing_opps:
        elements.append(Paragraph(", ".join([sanitize_text(x) for x in missing_opps]), normal_text))
    else:
        elements.append(Paragraph("All essential conversion paths are active.", normal_text))
    elements.append(Spacer(1, 5))
    
    elements.append(Paragraph("CTA Strength Breakdown", bold_text))
    cta_breakdown = f"Urgency: {report_data.get('cta_urgency_score', 0)}/10 | Placement: {report_data.get('cta_placement_quality', 'Suboptimal')} | Visibility: {report_data.get('cta_visibility_rating', 'Moderate')}"
    elements.append(Paragraph(cta_breakdown, normal_text))
    elements.append(Spacer(1, 10))
    
    elements.append(Paragraph("Conversion Intelligence Insight", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('conversion_intelligence_insight', '')), normal_text))
    elements.append(Spacer(1, 15))
    
    # Trust & Credibility Analysis
    elements.append(Paragraph("Trust Intelligence & Technical Security", subtitle_style))
    
    ha = report_data.get('has_analytics', {})
    trust_checklist_data = [
        [Paragraph("Google Analytics", normal_text), Paragraph("Yes" if ha.get('google_analytics') else "No", normal_text),
         Paragraph("Contact Form", normal_text), Paragraph("Yes" if report_data.get('has_lead_capture') else "No", normal_text)],
        [Paragraph("Facebook Pixel", normal_text), Paragraph("Yes" if ha.get('facebook_pixel') else "No", normal_text),
         Paragraph("Newsletter", normal_text), Paragraph("Yes" if report_data.get('has_newsletter') else "No", normal_text)],
        [Paragraph("SSL Valid", normal_text), Paragraph("Yes" if report_data.get('seo_ssl') else "No", normal_text),
         Paragraph("Title Tag", normal_text), Paragraph("Yes" if report_data.get('seo_title') else "No", normal_text)]
    ]
    tct = Table(trust_checklist_data, colWidths=[125, 125, 125, 125])
    tct.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(tct)
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("Credibility Impact Insight", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('credibility_impact_insight', '')), normal_text))
    elements.append(Spacer(1, 10))
    
    elements.append(Paragraph("Trust Recommendation", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('ai_trust_recommendation', '')), normal_text))
    elements.append(Spacer(1, 15))
    
    # Detailed SEO & Lighthouse
    elements.append(Paragraph("Google SEO Score & Technical Details", subtitle_style))
    
    lh_issues = report_data.get('lighthouse_issues', {})
    if lh_issues:
        issues_text = []
        for cat, issues in lh_issues.items():
            if issues:
                issues_text.append(f"<b>{str(cat).capitalize()}</b>: {', '.join(issues)}")
        if issues_text:
            elements.append(Paragraph("Critical Technical Issues Detected:", bold_text))
            for it in issues_text:
                elements.append(Paragraph(it, normal_text))
            elements.append(Spacer(1, 10))

    seo_details = f"<b>Live Page Load Speed:</b> {report_data.get('load_time', 'Unknown')}s | <b>Primary Tech Stack:</b> {report_data.get('tech_stack', 'Unknown')}"
    elements.append(Paragraph(seo_details, normal_text))
    elements.append(Spacer(1, 15))
    
    # AI Search Visibility (AEO)
    elements.append(Paragraph("AI Visibility (AEO) Analysis", subtitle_style))
    
    aeo_breakdown = f"<b>Brand Authority:</b> {min(100, aeo_score + 5)}% | <b>Conversational Ranking:</b> {max(0, aeo_score - 10)}% | <b>Topical Relevance:</b> {aeo_score}%"
    elements.append(Paragraph(aeo_breakdown, normal_text))
    elements.append(Spacer(1, 10))
    
    elements.append(Paragraph("AEO Probe Response", bold_text))
    quote_style = ParagraphStyle(
        'QuoteStyle', parent=normal_text,
        leftIndent=15, rightIndent=15,
        textColor=BRAND_DARK, fontName='Helvetica-Oblique'
    )
    aeo_quote = str(report_data.get('aeo_probe_response', 'No AI recognition data.'))[:500]
    elements.append(Paragraph(f'"{sanitize_text(aeo_quote)}"', quote_style))
    elements.append(Spacer(1, 15))
    
    # Recommendations & Strategy
    elements.append(Paragraph("Strategic Recommendations & Action Plan", subtitle_style))
    
    elements.append(Paragraph("Executive AI Recommendation", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('executive_ai_recommendation', '')), normal_text))
    elements.append(Spacer(1, 10))
    
    elements.append(Paragraph("Rebranding Pitch", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('rebranding_pitch', '')), normal_text))
    elements.append(Spacer(1, 15))
    
    # 1. Revenue Impact & Conversion Risk
    elements.append(Paragraph("Revenue Impact Forecast & Conversion Opportunity", subtitle_style))
    
    revenue_data = [
        [Paragraph("Revenue Leak Severity", bold_text), Paragraph(sanitize_text(report_data.get('revenue_leak_severity', 'Low')), normal_text)],
        [Paragraph("Monthly Revenue Impact", bold_text), Paragraph(f"${report_data.get('revenue_leak_amount', 0):,}/mo", normal_text)],
        [Paragraph("Annual Projected Loss", bold_text), Paragraph(f"${report_data.get('annual_opportunity_loss', report_data.get('revenue_leak_amount', 0) * 12):,}/yr", normal_text)],
        [Paragraph("Urgency", bold_text), Paragraph(sanitize_text(report_data.get('urgency_severity', '90+ Days')), normal_text)],
    ]
    rev_table = Table(revenue_data, colWidths=[150, 350])
    rev_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0,0), (0,-1), BRAND_LIGHT_BG),
        ('PADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(rev_table)
    elements.append(Spacer(1, 10))
    
    elements.append(Paragraph("Revenue Impact Insight", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('revenue_impact_insight', report_data.get('revenue_leak_explanation', ''))), normal_text))
    elements.append(Spacer(1, 15))

    # 2. Market Position & Momentum
    elements.append(Paragraph("Market Position & Competitor Momentum", subtitle_style))
    
    market_data = [
        [Paragraph("Industry Tier", bold_text), Paragraph(sanitize_text(report_data.get('industry_tier', 'Unknown')), normal_text)],
        [Paragraph("Industry Percentile", bold_text), Paragraph(f"{report_data.get('industry_percentile', 0)}%", normal_text)],
        [Paragraph("Lead Quality Score", bold_text), Paragraph(f"{report_data.get('lead_quality_score', 0)}%", normal_text)],
        [Paragraph("Buyer Intent Strength", bold_text), Paragraph(sanitize_text(report_data.get('buyer_intent_strength', 'Moderate')), normal_text)],
        [Paragraph("Commercial Readiness", bold_text), Paragraph(sanitize_text(report_data.get('commercial_readiness_maturity', 'Moderate')), normal_text)],
        [Paragraph("Momentum Score", bold_text), Paragraph(f"{report_data.get('momentum_score', 0)}/100", normal_text)],
        [Paragraph("Strategic Risk Level", bold_text), Paragraph(sanitize_text(report_data.get('strategic_risk_level', 'Moderate')), normal_text)],
    ]
    mkt_table = Table(market_data, colWidths=[150, 350])
    mkt_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0,0), (0,-1), BRAND_LIGHT_BG),
        ('PADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(mkt_table)
    elements.append(Spacer(1, 10))
    
    elements.append(Paragraph("Competitor Advantage Insight", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('keyword_visibility_gap_competitor_advantage', '')), normal_text))
    elements.append(Spacer(1, 15))
    
    # 3. Schema & Keyword Visibility Gaps
    elements.append(Paragraph("Schema & Search Visibility Gaps", subtitle_style))
    
    schema_data = [
        [Paragraph("Schema Coverage Score", bold_text), Paragraph(f"{report_data.get('schema_coverage_score', 0)}%", normal_text)],
        [Paragraph("Visibility Impact", bold_text), Paragraph(sanitize_text(report_data.get('schema_visibility_impact', 'Low')), normal_text)],
        [Paragraph("Keyword Gap Level", bold_text), Paragraph(sanitize_text(report_data.get('keyword_visibility_gap_level', 'Low')), normal_text)],
        [Paragraph("Search Impact", bold_text), Paragraph(sanitize_text(report_data.get('keyword_visibility_gap_search_impact', 'Low')), normal_text)],
    ]
    sch_table = Table(schema_data, colWidths=[150, 350])
    sch_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0,0), (0,-1), BRAND_LIGHT_BG),
        ('PADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(sch_table)
    elements.append(Spacer(1, 10))
    
    kw_opps = report_data.get('keyword_visibility_gap_opportunities', '')
    if kw_opps:
        elements.append(Paragraph("High-Intent Keyword Opportunities:", bold_text))
        elements.append(Paragraph(sanitize_text(kw_opps), normal_text))
        elements.append(Spacer(1, 5))
    
    elements.append(Paragraph("Schema AI Insight", bold_text))
    elements.append(Paragraph(sanitize_text(report_data.get('schema_gap_insight', '')), normal_text))
    elements.append(Spacer(1, 15))

    # 4. Strategic Action Plan
    elements.append(Paragraph("AI Strategic Action Plan Roadmap", subtitle_style))
    
    plan_steps = report_data.get('ai_strategic_plan', [])
    if plan_steps and isinstance(plan_steps, list):
        plan_data = [[Paragraph("Priority", bold_text), Paragraph("Action", bold_text), Paragraph("Impact", bold_text)]]
        for step in plan_steps:
            if isinstance(step, dict):
                plan_data.append([
                    Paragraph(sanitize_text(step.get('priority', '')), normal_text),
                    Paragraph(sanitize_text(step.get('action', '')), normal_text),
                    Paragraph(sanitize_text(step.get('impact', '')), normal_text)
                ])
        
        if len(plan_data) > 1:
            plan_table = Table(plan_data, colWidths=[80, 340, 80])
            plan_table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('BACKGROUND', (0,0), (-1,0), BRAND_LIGHT_BG),
                ('PADDING', (0,0), (-1,-1), 8),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ]))
            elements.append(plan_table)
        else:
            elements.append(Paragraph("No specific action steps generated.", normal_text))
    else:
        elements.append(Paragraph("No specific action steps generated.", normal_text))
        
    elements.append(Spacer(1, 20))
    
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        canvas.setFillColor(BRAND_GRAY)
        canvas.drawString(40, 20, f"Softree Technology Audit Report - {website}")
        # letter width is 612
        canvas.drawRightString(572, 20, f"Page {doc.page}")
        canvas.restoreState()
        
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_excel_report(report_data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Audit Report"
    
    ws['A1'] = "Softree AI Audit Report"
    ws['A2'] = "Website"
    ws['B2'] = report_data.get('website', '')
    
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    
    return excel_bytes
